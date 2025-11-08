import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from transaction import Journal


class Visualization:
    def __init__(self, transactions: list, journal: Journal):
        self.transactions = transactions
        self.journal = journal

    def _transactions_to_dataframe(self) -> pd.DataFrame:
        """
        Converts "Transaction" objects to DataFrame
        """
        data = []
        for txn in self.transactions:
            if txn.entry:
                data.append({
                    'Transaction ID': txn.entry.get('id'),
                    'Date': txn.date,
                    'Description': txn.description,
                    'Debit Category': txn.entry.get('debit_category'),
                    'Debit Account': txn.entry.get('debit_account'),
                    'Debit Amount': txn.entry.get('debit'),
                    'Credit Category': txn.entry.get('credit_category'),
                    'Credit Account': txn.entry.get('credit_account'),
                    'Credit Amount': txn.entry.get('credit')
                })
        return pd.DataFrame(data)

    def _build_adjusted_trial_balance_sheet(self, worksheet):
        """
        Build adjusted trial balance with proper formatting
        """
        account_balances = self.journal.get_account_balances()

        section_header_font = Font(bold=True, size=10, color='FFFFFF')
        section_header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        total_font = Font(bold=True)
        total_fill = PatternFill(start_color='D9E8F5', end_color='D9E8F5', fill_type='solid')
        data_font = Font(size=10)
        white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        row = 1

        # 1. Title: Company name
        title_cell = worksheet.cell(row=row, column=1)
        title_cell.value = self.journal.journal_name
        title_cell.font = Font(size=14, bold=True)
        title_cell.fill = white_fill
        title_cell.alignment = Alignment(horizontal='center', vertical='center')

        title_cell2 = worksheet.cell(row=row, column=2)
        title_cell2.fill = white_fill

        worksheet.merge_cells(f'A{row}:B{row}')
        row += 1

        # 2. Title: "Adjusted Trial Balance"
        subtitle_cell = worksheet.cell(row=row, column=1)
        subtitle_cell.value = "Adjusted Trial Balance"
        subtitle_cell.font = Font(size=14, bold=True)
        subtitle_cell.fill = white_fill
        subtitle_cell.alignment = Alignment(horizontal='center', vertical='center')

        subtitle_cell2 = worksheet.cell(row=row, column=2)
        subtitle_cell2.fill = white_fill

        worksheet.merge_cells(f'A{row}:B{row}')
        row += 1

        # 3. Title: "Operating period" start date to end date
        period_text = f"Operating period {self.journal.start_date} to {self.journal.end_date}"
        period_cell = worksheet.cell(row=row, column=1)
        period_cell.value = period_text
        period_cell.font = Font(size=14, bold=True)
        period_cell.fill = white_fill
        period_cell.alignment = Alignment(horizontal='center', vertical='center')

        period_cell2 = worksheet.cell(row=row, column=2)
        period_cell2.fill = white_fill

        worksheet.merge_cells(f'A{row}:B{row}')
        row += 1

        worksheet.column_dimensions['A'].width = 30
        worksheet.column_dimensions['B'].width = 15

        totals = {'asset': 0, 'liability': 0, 'equity': 0}

        # Assets
        if 'asset' in account_balances:
            accounts = account_balances['asset']

            header_cell1 = worksheet.cell(row=row, column=1)
            header_cell1.value = "Assets:"
            header_cell1.font = section_header_font
            header_cell1.fill = section_header_fill

            header_cell2 = worksheet.cell(row=row, column=2)
            header_cell2.fill = section_header_fill
            row += 1

            category_total = 0

            for account_name, balance in accounts.items():
                if balance == 0:
                    continue

                data_cell1 = worksheet.cell(row=row, column=1)
                data_cell1.value = account_name
                data_cell1.font = data_font
                data_cell1.fill = white_fill

                data_cell2 = worksheet.cell(row=row, column=2)
                data_cell2.value = abs(balance)
                data_cell2.font = data_font
                data_cell2.alignment = Alignment(horizontal='right')
                data_cell2.fill = white_fill

                category_total += balance
                row += 1

            totals['asset'] = category_total

            total_cell1 = worksheet.cell(row=row, column=1)
            total_cell1.value = "Total assets:"
            total_cell1.font = total_font
            total_cell1.fill = total_fill

            total_cell2 = worksheet.cell(row=row, column=2)
            total_cell2.value = abs(category_total)
            total_cell2.font = total_font
            total_cell2.fill = total_fill
            total_cell2.alignment = Alignment(horizontal='right')
            total_cell2.border = Border(
                bottom=Side(style='double'),
                top=Side(style='thin'),
                left=Side(style='thin'),
                right=Side(style='thin')
            )
            row += 2

        # Liabilities and equity
        header_cell1 = worksheet.cell(row=row, column=1)
        header_cell1.value = "Liabilities and Equity:"
        header_cell1.font = section_header_font
        header_cell1.fill = section_header_fill

        header_cell2 = worksheet.cell(row=row, column=2)
        header_cell2.fill = section_header_fill
        row += 1

        if 'liability' in account_balances:
            accounts = account_balances['liability']

            liability_total = 0

            for account_name, balance in accounts.items():
                if balance == 0:
                    continue

                data_cell1 = worksheet.cell(row=row, column=1)
                data_cell1.value = account_name
                data_cell1.font = data_font
                data_cell1.fill = white_fill

                data_cell2 = worksheet.cell(row=row, column=2)
                data_cell2.value = abs(balance)
                data_cell2.font = data_font
                data_cell2.alignment = Alignment(horizontal='right')
                data_cell2.fill = white_fill

                liability_total += balance
                row += 1

            totals['liability'] = liability_total

            total_cell1 = worksheet.cell(row=row, column=1)
            total_cell1.value = "Total liabilities:"
            total_cell1.font = total_font
            total_cell1.fill = total_fill

            total_cell2 = worksheet.cell(row=row, column=2)
            total_cell2.value = abs(liability_total)
            total_cell2.font = total_font
            total_cell2.fill = total_fill
            total_cell2.alignment = Alignment(horizontal='right')
            total_cell2.border = Border(
                bottom=Side(style='thin'),
                top=Side(style='thin'),
                left=Side(style='thin'),
                right=Side(style='thin')
            )
            total_cell2.fill = total_fill
            row += 1

        if 'equity' in account_balances:
            accounts = account_balances['equity']

            equity_total = 0

            for account_name, balance in accounts.items():
                if balance == 0:
                    continue

                data_cell1 = worksheet.cell(row=row, column=1)
                data_cell1.value = account_name
                data_cell1.font = data_font
                data_cell1.fill = white_fill

                data_cell2 = worksheet.cell(row=row, column=2)
                data_cell2.value = abs(balance)
                data_cell2.font = data_font
                data_cell2.alignment = Alignment(horizontal='right')
                data_cell2.fill = white_fill

                equity_total += balance
                row += 1

            totals['equity'] = equity_total

            total_cell1 = worksheet.cell(row=row, column=1)
            total_cell1.value = "Total equity:"
            total_cell1.font = total_font
            total_cell1.fill = total_fill

            total_cell2 = worksheet.cell(row=row, column=2)
            total_cell2.value = abs(equity_total)
            total_cell2.font = total_font
            total_cell2.fill = total_fill
            total_cell2.alignment = Alignment(horizontal='right')
            total_cell2.border = Border(
                bottom=Side(style='double'),
                top=Side(style='thin'),
                left=Side(style='thin'),
                right=Side(style='thin')
            )
            row += 2

        # Total liabilities and equity
        total_liability_equity = totals['liability'] + totals['equity']

        total_cell1 = worksheet.cell(row=row, column=1)
        total_cell1.value = "Total liabilities and equity:"
        total_cell1.font = total_font
        total_cell1.fill = total_fill

        total_cell2 = worksheet.cell(row=row, column=2)
        total_cell2.value = abs(total_liability_equity)
        total_cell2.font = total_font
        total_cell2.fill = total_fill
        total_cell2.alignment = Alignment(horizontal='right')
        total_cell2.border = Border(
            bottom=Side(style='double'),
            top=Side(style='double'),
            left=Side(style='thin'),
            right=Side(style='thin')
        )

    def export_to_excel(self, filename: str):
        """
        Export transactions to first worksheet and trial balance to second worksheet in Excel file.
        """
        transactions = self._transactions_to_dataframe()

        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            transactions.to_excel(writer, sheet_name='Transactions', index=False)

            workbook = writer.book
            worksheet1 = writer.sheets['Transactions']
            worksheet2 = workbook.create_sheet('Trial Balance')

            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

            for col_num, value in enumerate(transactions.columns.values):
                cell = worksheet1.cell(row=1, column=col_num + 1)
                cell.font = header_font
                cell.fill = header_fill

            for row in worksheet1.iter_rows(min_row=2, max_row=worksheet1.max_row, min_col=1, max_col=len(transactions.columns)):
                for cell in row:
                    cell.fill = white_fill

            for col in worksheet1.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet1.column_dimensions[column].width = adjusted_width

            self._build_adjusted_trial_balance_sheet(worksheet2)
